
from datetime import datetime
start_time = datetime.now()
import os

#Help
def octant_analysis(mod=5000):
	octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
	import pandas as pd
	try:
		inputfiles=os.listdir("input")
		for reading in inputfiles:
			#outdata=[pd.read_csv("output\\"+i+"vel_octant_analysis_mod"+str(mod))]
			data=pd.read_excel("input\\"+reading)
			lastval=data.index[-1]
			mean1=data["U"].mean()
			mean2=data["V"].mean()
			mean3=data["W"].mean()
			data["U Avg"]=""
			data["V Avg"]=""
			data["W Avg"]=""
			data["U'=U-U Avg"]=""
			data["V'=V-V Avg"]=""
			data["W'=W-W Avg"]=""
			data["Octant"]=""
			rm1=format(mean1,".3f")
			rm2=format(mean2,".3f")
			rm3=format(mean3,".3f")
			data.at[0,"U Avg"]=rm1
			data.at[0,"V Avg"]=rm2
			data.at[0,"W Avg"]=rm3
			for i in range(0,lastval+1,1):
				a=round(data["U"][i],3)-float(data["U Avg"][0])
				data.at[i,"U'=U-U Avg"]=format(a,".3f")
			for i in range(0,lastval+1,1):
				a=round(data["V"][i],3)-float(data["V Avg"][0])
				data.at[i,"V'=V-V Avg"]=format(a,".3f")
			for i in range(0,lastval+1,1):
				a=round(data["W"][i],3)-float(data["W Avg"][0])
				data.at[i,"W'=W-W Avg"]=format(a,".3f")	
			for i in range(0,lastval+1,1):
				#print(type(data["U'=U-U Avg"][i]))
				u=float(data["U'=U-U Avg"][i])
				v=float(data["V'=V-V Avg"][i])
				w=float(data["W'=W-W Avg"][i])
				
				if((u==0) or(v==0)or(w==0)):
					data.at[i,'Octant']=1
				else:	
					if u>0:
						if v>0:
							if w>0:
								
								data.at[i,'Octant']=1
							else:
								
								data.at[i,'Octant']=-1
						else:
							if w>0:
								
								data.at[i,'Octant']=4
							else:
								
								data.at[i,'Octant']=-4
					else:
						if v>0:
							if w>0:
								data.at[i,'Octant']=2
							else:
								data.at[i,'Octant']=-2
						else:
							if w>0:
								data.at[i,'Octant']=3
							else: 
								data.at[i,'Octant']=-3				
			data[""]=""
			data[" "]=""
			data["Overall Octant Count"]=""
			data.at[1,"Overall Octant Count"]="Octant ID"
			data.at[2,"Overall Octant Count"]="Overall Count"
			data.at[2," "]="Mod "+str(mod)
			data["  "]=""
			data["   "]=""
			data["    "]=""
			data["     "]=""
			data["      "]=""
			data["       "]=""
			data["        "]=""
			data["         "]=""
			data.at[1,"  "]="1"
			data.at[1,"   "]="-1"
			data.at[1,"    "]="2"
			data.at[1,"     "]="-2"
			data.at[1,"      "]="3"
			data.at[1,"       "]="-3"
			data.at[1,"        "]="4"
			data.at[1,"         "]="-4"
			for i in range(10,22,1):
				data[" "*i]=""	
			data.at[1," "*18]="Rank1 Octant ID"
			data.at[1," "*19]="Rank1 Octant Name"
			values=[1,-1,2,-2,3,-3,4,-4]
			for i in range(0,8,1):
				x=" "*(i+10)
				data.at[1,x]="Rank Octant "+str(values[i])
			
			#s=1
			#for i in values:
			#	data.at[0,str(i)]="Rank"+" "+str(s)
			#	s+=1
			#data.at[0,"1"]="Rank 1"
			#lastval=data.index[-1]
			
			count1=0
			countm1=0
			count2=0
			countm2=0
			count3=0
			countm3=0
			count4=0
			countm4=0
			multiply=lastval//mod
			counterforeachoctant=2
			a=0
			b=0
			c=0
			d=0
			e=0
			f=0
			g=0
			h=0
			for j in range(1,multiply+2,1):
				if(j==1):
					data.at[j+2,"Overall Octant Count"]="0000"+"-"+str(mod-1)
				else:
					if(j<=multiply):
						num=str(mod*(j-1))+"-"+str(mod*j - 1)
						data.at[j+2,"Overall Octant Count"]=num
					else:
						num=str(mod*(j-1))+"-"+str(lastval)
						data.at[j+2,"Overall Octant Count"]=num
			rank1count=[0,0,0,0,0,0,0,0]
			for i in range(0,lastval+1,1):
				u=data["U'=U-U Avg"][i]
				v=data["V'=V-V Avg"][i]
				w=data["W'=W-W Avg"][i]                   
				if((data["Octant"][i])==1):
					count1+=1
				elif((data["Octant"][i])==-1):
					countm1+=1
				elif((data["Octant"][i])==2):
					count2+=1
				elif((data["Octant"][i])==-2):
					countm2+=1   
				elif((data["Octant"][i])==3):
					count3+=1
				elif((data["Octant"][i])==-3):
					countm3+=1
				elif((data["Octant"][i])==4):
					count4+=1
				elif((data["Octant"][i])==-4):
					countm4+=1
				if(((i!=0)and((i+1)%mod==0))or(i==lastval)):
					data.at[counterforeachoctant+1,"  "]=count1-a
					data.at[counterforeachoctant+1,"   "]=countm1-b
					data.at[counterforeachoctant+1,"    "]=count2-c
					data.at[counterforeachoctant+1,"     "]=countm2-d
					data.at[counterforeachoctant+1,"      "]=count3-e
					data.at[counterforeachoctant+1,"       "]=countm3-f
					data.at[counterforeachoctant+1,"        "]=count4-g
					data.at[counterforeachoctant+1,"         "]=countm4-h
					counts=[count1-a,countm1-b,count2-c,countm2-d,count3-e,countm3-f,count4-g,countm4-h]
					ncounts=sorted(counts)
					for i in range(0,8,1):
						for j in range(0,8,1):
							if (ncounts[i]==counts[j]):
								data.at[counterforeachoctant+1," "*(j+10)]=8-i
								data.at[counterforeachoctant+1," "*18]=values[j]
								data.at[counterforeachoctant+1," "*19]=octant_name_id_mapping[str(values[j])]
								if (i==7):
									rank1count[j]=rank1count[j]+1
					a,b,c,d,e,f,g,h=count1,countm1,count2,countm2,count3,countm3,count4,countm4
					counterforeachoctant+=1
				else:
					continue
				
			data.at[2,"  "]=count1
			data.at[2,"   "]=countm1
			data.at[2,"    "]=count2
			data.at[2,"     "]=countm2
			data.at[2,"      "]=count3
			data.at[2,"       "]=countm3
			data.at[2,"        "]=count4
			data.at[2,"         "]=countm4
			
			ovcounts=[count1,countm1,count2,countm2,count3,countm3,count4,countm4]    
			novcounts=sorted(ovcounts)
			for i in range(0,8,1):
				for j in range(0,8,1):
					if (novcounts[i]==ovcounts[j]):
						data.at[2," "*(j+10)]=8-i
						data.at[2," "*18]=values[j]
						data.at[2," "*19]=octant_name_id_mapping[str(values[j])]
			newpos=5+(lastval//mod)
			data.at[newpos," "*16]="Octant ID"
			data.at[newpos," "*17]="Octant Name"
			data.at[newpos," "*18]="Count of Rank 1 Mod Values"
			for i in range(0,8,1):
				data.at[newpos+1+i," "*16]=values[i]
				data.at[newpos+1+i," "*17]=octant_name_id_mapping[str(values[i])]
				data.at[newpos+1+i," "*18]=rank1count[i]

			
			#################################
			#This is the overall transition count
			data["Overall Transition Count"]=""
			for i in range(23,32,1):
				data[" "*i]=""
			data.at[2," "*21]="From"	
			data.at[0," "*23]="To"
			data.at[1,"Overall Transition Count"]="Octant"	
			listofvalues=[+1,-1,+2,-2,+3,-3,+4,-4]
			for i in range(0,8,1):
				data.at[i+2,"Overall Transition Count"]=listofvalues[i]
				data.at[1," "*(i+23)]=listofvalues[i]             
			##################################

			
			#####################################
			#To count the total octant transitions
			#We iterate through all the values in the dataframe in the "octant" column
			#if there is a change in the value from ith to (i+1)th octant value then the position of writing 
			#in the excel sheet is stored in the two variables "toposition" and "fromposition" to uniquely
			#write to a particular cell in the excel sheet
			for i in range(0,lastval,1):
				values=[1,-1,2,-2,3,-3,4,-4]
				for j in range(0,8,1):
					if(values[j]==data["Octant"][i]):
						fromposition=2+j
					if(values[j]==data["Octant"][i+1]):
						toposition=" "*(23+j)
				if(not(data[toposition][fromposition])):
					data.at[fromposition,toposition]=1 
				else:    
					data.at[fromposition,toposition]=data[toposition][fromposition]+1
			######################################
				
			
			multiply = lastval//mod
			#We initialize all the octant transitions in each mod range by filling the matrix with zeroes
			#the variable k helps to get to the beginning of a new table of writing cells
			#the exact position of the writing cells is determined by two variables "toposition" and "fromposition"
			for k in range(0,multiply+1,1):
				for i in range(0,8,1):
					for j in range(0,8,1):
						fromposition=16+(13*k)+i
						toposition=" "*(23+j)
						data.at[fromposition,toposition]=0 


			
			##########################################3
			#Writing down the table headers and other values before writing the transition counts
			for k in range(0,multiply+1,1):
				data.at[13+(13*k),"Overall Transition Count"]="Mod Transition Count"
				data.at[14+(13*k)," "*23]="To"
				data.at[15+(13*k),"Overall Transition Count"]="Octant"
				for i in range(0,8,1):
					data.at[16+(13*k)+i,"Overall Transition Count"]=values[i]
					data.at[15+(13*k)," "*(i+23)]=values[i]   
				data.at[16+(13*k)," "*21]="From"
				if(k==0):
					data.at[14+(13*k),"Overall Transition Count"]="0000"+"-"+str(mod-1)
				else:
					if(k<multiply):
						num=str(mod*(k))+"-"+str(mod*(k+1) - 1)
						data.at[14+(13*k),"Overall Transition Count"]=num
					else:
						num=str(mod*(k))+"-"+str(lastval)
						data.at[14+(13*k),"Overall Transition Count"]=num   
				
				for i in range(0,8,1):
					data.at[i+2,"Overall Transition Count"]=listofvalues[i]
					data.at[1," "*(i+23)]=listofvalues[i]   
				#######################################

				
				for i in range(mod*(k),mod*(k+1),1):
					if(i==lastval):
						break
					for j in range(0,8,1):
						if(values[j]==data["Octant"][i]):
							fromposition=16+(13*k)+j
						if(values[j]==data["Octant"][i+1]):
							toposition=" "*(23+j)
					if(not(data[toposition][fromposition])):
						data.at[fromposition,toposition]=1 #inserts 1 if value is empty
					else:    
						data.at[fromposition,toposition]=data[toposition][fromposition]+1   #adds 1 if value is not empty          
			data["Longest Subsequence Length"]=""
			data.at[1,"Longest Subsequence Length"]="Octant"
			for i in range(0,8,1):
				data.at[2+i,"Longest Subsequence Length"]=values[i]
			data[" "*33]=""
			data[" "*34]=""
			data[" "*35]=""
			
			
			##############################
			count=[0,0,0,0,0,0,0,0] #creating an empty array to store the counted values of consecutive occurrences
			large=[0,0,0,0,0,0,0,0] #creating another empty array to store the largest consecutively counted occurrences
			for i in range(0,lastval,1):
				if(data['Octant'][i]==data['Octant'][i+1]):
					for j in range(0,8,1):
						if (data['Octant'][i]==values[j]):
							count[j]+=1
							pos=j
				else:
					if (count[pos]>large[pos]):
						large[pos]=count[pos]   
						count[pos]=0
						#print(i)
					else:
						large[pos]=large[pos]   
						count[pos]=0          
						
			for j in range(0,8,1):
				for i in range(0,lastval,1):
					if (data['Octant'][i]==values[j]):
						large[j]+=1
						break
			
			#code to count how many largest occurrences present for each octant value
			count2=[0,0,0,0,0,0,0,0]              
			for j in range(0,8,1):                   
				for i in range(0,lastval-large[j]+1,1):
					supercount=0
					for k in range(0,large[j],1):
						if(data['Octant'][i+k]==values[j]):
							supercount+=1
					if(supercount==large[j]):
						count2[j]+=1
			#code to insert the final values to the excel sheet
			for i in range(0,8,1):
				data.at[i+2," "*33]=large[i]
				data.at[1," "*33]="Longest Subsequence Length"
				data.at[i+2," "*34]=count2[i]
				data.at[1," "*34]="Count"
			data["Longest Subsequence Length with Range"]=""
			data[" "*37]=""	
			data[" "*38]=""

			timefromtopos=3
			for i in range(0,8,1):

				#Here we write the octant values as headings for each block in the excel sheet
				#We also add the headings "Time", "To" and "From" to the blocks for each octant
				data.at[timefromtopos-1,"Longest Subsequence Length with Range"]=values[i]
				data.at[timefromtopos,"Longest Subsequence Length with Range"]="Time"
				data.at[timefromtopos-1," "*37]=large[i]
				data.at[timefromtopos," "*37]="From"
				data.at[timefromtopos," "*38]="To"
				data.at[timefromtopos-1," "*38]=count2[i]
				
				anotherpos=1
				for j in range(0,lastval,1):
					newcount=1
					if(values[i]==data["Octant"][j]):   #this code checks the initial and final timestamp of the largest consecutive occerrence of an octant value
						for k in range(1,large[i],1):   #it goes to the octant value, checks if the next 'n' consecutive numbers are same as the current number
							if(large[i]<=(lastval-j)):  #where 'n' is the largest consecutive occerrence calculated previously
								if(data["Octant"][j]==data["Octant"][j+k]):
									newcount+=1
					if(newcount==large[i]):
							#writes the timestamp to the excel sheet
							data.at[timefromtopos+anotherpos," "*37]=data["T"][j]
							data.at[timefromtopos+anotherpos," "*38]=data["T"][j+large[i]-1]
							anotherpos+=1  
				timefromtopos+=2+count2[i]
			data.at[1,"Longest Subsequence Length with Range"]="Octant"	
			data.at[1," "*37]="Longest Subsequence Length"
			data.at[1," "*38]="Count"
			a=reading[0:-5]
			writer = pd.ExcelWriter("output\\"+a+"_vel_octant_analysis_mod"+str(mod)+".xlsx") 
			data.to_excel(writer, sheet_name='Sheet1', index=False, na_rep='NaN')
			for column in data:
				if (not (column in ["  ","   ","    ","     ","      ","       ","        ","         "," "*21," "*23," "*24," "*25," "*26," "*27," "*28," "*29," "*30," "*34," "*38])):
					width = max(data[column].astype(str).map(len).max(), len(column))
					c = data.columns.get_loc(column)
					writer.sheets['Sheet1'].set_column(c, c, width)
			writer.save()
			
			import openpyxl
			from openpyxl.styles import PatternFill
			from openpyxl.styles.borders import Border, Side
			#FFFF00 hex code for yellow used in the sample output
			a=reading[0:-5]
			book = openpyxl.load_workbook("output\\"+a+"_vel_octant_analysis_mod"+str(mod)+".xlsx")
			sheet = book["Sheet1"]
			
			'''border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
			sheet.cell(row=3, column=2).border = border'''
			
			alpha=['W','X','Y','Z','AA','AB','AC','AD']
			for i in range(0,8,1):
				j=2
				while (data[" "*(i+10)][j]):
					if(data[" "*(i+10)][j]==1):
						fill = PatternFill(patternType='solid', fgColor='FFFF00')
						sheet[alpha[i]+str(j+2)].fill = fill
					j+=1
			alpha2=["AJ","AK","AL","AM","AN","AO","AP","AQ"]		
			for i in range(0,8,1):
				lis=[]
				for j in range(0,8,1):
					lis.append(data[" "*(j+23)][i+2])	
				maxm=max(lis)
				for j in range(0,8,1):
					if(data[" "*(j+23)][i+2]==maxm):
						fill = PatternFill(patternType='solid', fgColor='FFFF00')
						sheet[alpha2[j]+str(i+4)].fill = fill	

			for k in range(0,multiply+1,1):
				for i in range(0,8,1):   
					lis2=[]
					for j in range(0,8,1):
						lis2.append(data[" "*(j+23)][16+(13*k)+i])
					maxm2=max(lis2)
					for j in range(0,8,1):
						if(data[" "*(j+23)][16+(13*k)+i]==maxm2):
							fill = PatternFill(patternType='solid', fgColor='FFFF00')
							sheet[alpha2[j]+str(16+(13*k)+i+2)].fill = fill

			
			def cell_border(ws, cell_range):
				rows = ws[cell_range]
				side = Side(border_style='thin', color="FF000000")
				rows = list(rows)    
				for y, cells in enumerate(rows):
					for x, cell in enumerate(cells):
						border = Border(left=cell.border.left,right=cell.border.right,top=cell.border.top,bottom=cell.border.bottom)
						border.left = side
						border.right = side
						border.top = side
						border.bottom = side
						cell.border = border
			cell_border(sheet,"N3:"+"AF"+str(5+multiply))
			cell_border(sheet,"AI3:AQ11")
			cell_border(sheet,"AC10:AE18")
			cell_border(sheet,"AS3:AU11")
			lastbound=0
			cell_border(sheet,"AW3:AY"+str(16+sum(count2)+3))
			for k in range(0,multiply+1,1):
				cell_border(sheet,"AI"+str(17+k*13)+":AQ"+str(17+k*13+8))
			a=reading[0:-5]
			book.save("output\\"+a+"_vel_octant_analysis_mod"+str(mod)+".xlsx")
	except FileNotFoundError:
		print("Incorrect file name")  
	else:
		pass
		#write some code


from platform import python_version
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


mod=5000
octant_analysis(mod)






#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
